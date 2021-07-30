#!/usr/bin/env python
## Copyright (c) 2019, Alliance for Open Media. All rights reserved
##
## This source code is subject to the terms of the BSD 2 Clause License and
## the Alliance for Open Media Patent License 1.0. If the BSD 2 Clause License
## was not distributed with this source code in the LICENSE file, you can
## obtain it at www.aomedia.org/license/software. If the Alliance for Open
## Media Patent License 1.0 was not distributed with this source code in the
## PATENTS file, you can obtain it at www.aomedia.org/license/patent.
##
__author__ = "maggie.sun@intel.com, ryanlei@fb.com"

import re
import openpyxl
import xlsxwriter
import shutil
import Config
from Config import QPs, DnScaleRatio, CTC_ASXLSTemplate, CTC_RegularXLSTemplate, InterpolatePieces, \
    UsePCHIPInterpolation
import Utils
from Utils import ParseCSVFile, plot_rd_curve, Interpolate_Bilinear, Interpolate_PCHIP, convex_hull
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from CalcBDRate import BD_RATE

qtys = ["psnr_y", "psnr_u", "psnr_v", "overall_psnr", "ssim_y", "ms_ssim_y",
        "vmaf", "vmaf_neg", "psnr_hvs","ciede2k", "apsnr_y", "apsnr_u",
        "apsnr_v", "overall_apsnr"]

'''
csv_files = {
    "HM_CloseGOP":
    {
        "LD":     "D:\\HEVC-AV1-Study\\HEVC-RA-CloseGop\\analysis\\rdresult\\RDResults_hm_hevc_RA_Preset_0.csv",
    },
    "HM_OpenGOP":
    {
        "LD":     "D:\\HEVC-AV1-Study\\HEVC-RA-OpenGop\\analysis\\rdresult\\RDResults_hm_hevc_RA_Preset_0.csv",
    },
    "AV1_CloseGOP":
    {
        "LD":     "D:\\HEVC-AV1-Study\\AV1-RA-CloseGop\\analysis\\rdresult\\RDResults_aom_av1_RA_Preset_0.csv",
    },
    "AV1_OpenGOP":
    {
        "LD":     "D:\\HEVC-AV1-Study\\AV1-RA-OpenGop\\analysis\\rdresult\\RDResults_aom_av1_RA_Preset_0.csv",
    },
}
'''
csv_files = {
    "v1.0.0":
    {
        "AI":     "D:\\AV2-CTC\\AV2-CTC-v1.0.0-Final\\analysis\\rdresult\\RDResults_aom_av2_AI_Preset_0.csv",
        "LD":     "D:\\AV2-CTC\\AV2-CTC-v1.0.0-Final\\analysis\\rdresult\\RDResults_aom_av2_LD_Preset_0.csv",
        "RA":     "D:\\AV2-CTC\\AV2-CTC-v1.0.0-Final\\analysis\\rdresult\\RDResults_aom_av2_RA_Preset_0.csv",
        "Still":  "D:\\AV2-CTC\\AV2-CTC-v1.0.0-Final\\analysis\\rdresult\\RDResults_aom_av2_STILL_Preset_0.csv",
        "AS":     "D:\\AV2-CTC\\AV2-CTC-v1.0.0-Final\\analysis\\rdresult\\RDResults_aom_av2_AS_Preset_0.csv",
    },
    "v1.0.1":
    {
        "AI":     "D:\\AV2-CTC\\AV2-CTC-v1.0.1\\analysis\\rdresult\\RDResults_aom_av2_AI_Preset_0.csv",
        "LD":     "D:\\AV2-CTC\\AV2-CTC-v1.0.1\\analysis\\rdresult\\RDResults_aom_av2_LD_Preset_0.csv",
        "RA":     "D:\\AV2-CTC\\AV2-CTC-v1.0.1\\analysis\\rdresult\\RDResults_aom_av2_RA_Preset_0.csv",
        "Still":  "D:\\AV2-CTC\\AV2-CTC-v1.0.1\\analysis\\rdresult\\RDResults_aom_av2_STILL_Preset_0.csv",
        "AS":     "D:\\AV2-CTC\\AV2-CTC-v1.0.1\\analysis\\rdresult\\RDResults_aom_av2_AS_Preset_0.csv",
    },
    "B034":
    {
        "AI":     "D:\\AV2-CTC\\AV2-CTC-B034\\analysis\\rdresult\\RDResults_aom_av2_AI_Preset_0.csv",
        "LD":     "D:\\AV2-CTC\\AV2-CTC-B034\\analysis\\rdresult\\RDResults_aom_av2_LD_Preset_0.csv",
        "RA":     "D:\\AV2-CTC\\AV2-CTC-B034\\analysis\\rdresult\\RDResults_aom_av2_RA_Preset_0.csv",
        "Still":  "D:\\AV2-CTC\\AV2-CTC-B034\\analysis\\rdresult\\RDResults_aom_av2_STILL_Preset_0.csv",
        "AS":     "D:\\AV2-CTC\\AV2-CTC-B034\\analysis\\rdresult\\RDResults_aom_av2_AS_Preset_0.csv",
    },
    "ext-quant":
    {
        "AI":     "D:\\AV2-CTC\\AV2-CTC-ExtQuant\\analysis\\rdresult\\RDResults_aom_av2_AI_Preset_0.csv",
        "LD":     "D:\\AV2-CTC\\AV2-CTC-ExtQuant\\analysis\\rdresult\\RDResults_aom_av2_LD_Preset_0.csv",
        "RA":     "D:\\AV2-CTC\\AV2-CTC-ExtQuant\\analysis\\rdresult\\RDResults_aom_av2_RA_Preset_0.csv",
        "Still":  "D:\\AV2-CTC\\AV2-CTC-ExtQuant\\analysis\\rdresult\\RDResults_aom_av2_STILL_Preset_0.csv",
        "AS":     "D:\\AV2-CTC\\AV2-CTC-ExtQuant\\analysis\\rdresult\\RDResults_aom_av2_AS_Preset_0.csv",
    },
    "sdp-off":
    {
        "AI":     "D:\\AV2-CTC\\AV2-CTC-SDP-OFF\\analysis\\rdresult\\RDResults_aom_av2_AI_Preset_0.csv",
        "LD":     "D:\\AV2-CTC\\AV2-CTC-SDP-OFF\\analysis\\rdresult\\RDResults_aom_av2_LD_Preset_0.csv",
        "RA":     "D:\\AV2-CTC\\AV2-CTC-SDP-OFF\\analysis\\rdresult\\RDResults_aom_av2_RA_Preset_0.csv",
        "Still":  "D:\\AV2-CTC\\AV2-CTC-SDP-OFF\\analysis\\rdresult\\RDResults_aom_av2_STILL_Preset_0.csv",
        "AS":     "D:\\AV2-CTC\\AV2-CTC-SDP-OFF\\analysis\\rdresult\\RDResults_aom_av2_AS_Preset_0.csv",
    },
    "sdp-on":
    {
        "AI":     "D:\\AV2-CTC\\AV2-CTC-SDP-ON\\analysis\\rdresult\\RDResults_aom_av2_AI_Preset_0.csv",
        "LD":     "D:\\AV2-CTC\\AV2-CTC-SDP-ON\\analysis\\rdresult\\RDResults_aom_av2_LD_Preset_0.csv",
        "RA":     "D:\\AV2-CTC\\AV2-CTC-SDP-ON\\analysis\\rdresult\\RDResults_aom_av2_RA_Preset_0.csv",
        "Still":  "D:\\AV2-CTC\\AV2-CTC-SDP-ON\\analysis\\rdresult\\RDResults_aom_av2_STILL_Preset_0.csv",
        "AS":     "D:\\AV2-CTC\\AV2-CTC-SDP-ON\\analysis\\rdresult\\RDResults_aom_av2_AS_Preset_0.csv",
    },
}

start_row = {
    "AI": 2,
    "AS": 2,
    "RA": 2,
    "Still": 2,
    "LD": 50,
}

formats = {
    "v1.0.0":       ['r', '-', 'o'],
    "v1.0.1":       ['g', '-', '*'],
    "B034":         ['k', '-', '^'],
    "ext-quant":    ['r', '-', '*'],
    "sdp-off":      ['b', '-', '+'],
    "sdp-on":       ['r', '-', '<'],
    "HM_CloseGOP":     ['r', '-', 'o'],
    "HM_OpenGOP":      ['b', '-', '+'],
    "AV1_CloseGOP":     ['g', '-', '>'],
    "AV1_OpenGOP":      ['k', '-', '*'],
}

AS_formats = {
    "3840x2160": ['r', '-.', 'o'],
    "2560x1440": ['g', '-.', '*'],
    "1920x1080": ['b', '-.', '^'],
    "1280x720":  ['y', '-.', '+'],
    "960x540":   ['c', '-.', 'x'],
    "640x360":   ['k', '-.', '<'],
}

anchor = "v1.0.0"
rd_curve_pdf = "rdcurve.pdf"

def WriteSheet(csv_file, sht, start_row):
    csv = open(csv_file, 'rt')
    row = start_row
    for line in csv:
        if not line.startswith('TestCfg'):
            words = re.split(',', line.strip())
            col = 1
            for word in words:
                mycell = sht.cell(row=row, column=col)
                if col >= 12 and col <= 30 and word != "":
                    mycell.value = float(word)
                else:
                    mycell.value = word
                col += 1
            row += 1
    csv.close()


def FillXlsFile():
    for tag in csv_files.keys():
        if tag == anchor:
            continue
        else:
            for cfg in csv_files[anchor].keys():
                anchor_sht_name = "Anchor-%s" % cfg
                test_sht_name = "Test-%s" % cfg
                if cfg == "AS":
                    xls_template = CTC_ASXLSTemplate
                    xls_file = "CTC_AS_%s-%s.xlsm" % (anchor, tag)
                    shutil.copyfile(xls_template, xls_file)
                    anchor_sht_name = "Anchor"
                    test_sht_name = "Test"
                elif cfg == "AI":
                    xls_template = CTC_RegularXLSTemplate
                    xls_file = "CTC_Regular_%s-%s.xlsm" % (anchor, tag)
                    shutil.copyfile(xls_template, xls_file)

                wb = openpyxl.load_workbook(filename=xls_file, read_only=False, keep_vba=True)
                anchor_sht = wb[anchor_sht_name]
                anchor_csv = csv_files[anchor][cfg]
                WriteSheet(anchor_csv, anchor_sht, start_row[cfg])

                test_sht = wb[test_sht_name]
                test_csv = csv_files[tag][cfg]
                WriteSheet(test_csv, test_sht, start_row[cfg])

                wb.save(xls_file)

def DrawRDCurve(records, anchor, pdf):
    with PdfPages(pdf) as export_pdf:
        for cfg in records[anchor].keys():
            videos = records[anchor][cfg].keys()
            for video in videos:
                if cfg == "AS":
                    DnScaledRes = [(int(3840/ratio), int(2160/ratio)) for ratio in DnScaleRatio]
                    Int_RDPoints = {}
                    # draw individual rd curves
                    for tag in records.keys():
                        Int_RDPoints[tag] = []
                        record = records[tag][cfg][video]
                        plt.figure(figsize=(15, 10))
                        plt.suptitle("%s : %s: %s" % (cfg, video, tag))

                        br = {}; apsnr = {}
                        for key in record.keys():
                            res = re.split('_', key)[0]
                            if res not in br.keys():
                                br[res] = []
                                apsnr[res] = []
                            br[res].append(record[key].bitrate)
                            apsnr[res].append(record[key].overall_apsnr)

                        for res in br.keys():
                            rdpnts = [(brt, qty) for brt, qty in zip(br[res], apsnr[res])]
                            if UsePCHIPInterpolation:
                                int_rdpnts = Interpolate_PCHIP(rdpnts, QPs['AS'][:], InterpolatePieces, True)
                            else:
                                int_rdpnts = Interpolate_Bilinear(rdpnts, QPs['AS'][:], InterpolatePieces, True)
                            Int_RDPoints[tag] += int_rdpnts
                            plot_rd_curve(br[res], apsnr[res], "overall_apsnr", res, "bitrate(Kbps)",
                                          AS_formats[res][0], AS_formats[res][1], AS_formats[res][2])
                        plt.legend()
                        plt.grid(True)
                        export_pdf.savefig()
                        plt.close()

                    #draw convex hull
                    plt.figure(figsize=(15, 10))
                    plt.suptitle("%s : %s: convex hull" % (cfg, video))
                    for tag in records.keys():
                        lower, upper = convex_hull(Int_RDPoints[tag])
                        br    = [h[0] for h in upper]
                        apsnr = [h[1] for h in upper]
                        plot_rd_curve(br, apsnr, "overall_apsnr", tag, "bitrate(Kbps)",
                                      formats[tag][0], formats[tag][1], formats[tag][2])
                    plt.legend()
                    plt.grid(True)
                    export_pdf.savefig()
                    plt.close()
                else:
                    # bit rate to quality
                    plt.figure(figsize=(15, 10))
                    plt.suptitle("%s : %s" % (cfg, video))
                    for tag in records.keys():
                        record = records[tag][cfg][video]
                        br    = [record[key].bitrate for key in record.keys()]
                        apsnr = [record[key].overall_apsnr for key in record.keys()]
                        plot_rd_curve(br, apsnr, "overall_apsnr", tag, "bitrate(Kbps)",
                                      formats[tag][0], formats[tag][1], formats[tag][2])
                    plt.legend()
                    plt.grid(True)
                    export_pdf.savefig()
                    plt.close()

def GetQty(record, qty):
    qtys = []
    for key in record.keys():
        if qty == 'psnr_y':
            qtys.append(record[key].psnr_y)
        elif qty == 'psnr_u':
            qtys.append(record[key].psnr_u)
        elif qty == 'psnr_v':
            qtys.append(record[key].psnr_v)
        elif qty == 'overall_psnr':
            qtys.append(record[key].overall_psnr)
        elif qty == 'ssim_y':
            qtys.append(record[key].ssim_y)
        elif qty == 'ms_ssim_y':
            qtys.append(record[key].ms_ssim_y)
        elif qty == 'vmaf':
            qtys.append(record[key].vmaf_y)
        elif qty == 'vmaf_neg':
            qtys.append(record[key].vmaf_y_neg)
        elif qty == 'psnr_hvs':
            qtys.append(record[key].psnr_hvs)
        elif qty == 'ciede2k':
            qtys.append(record[key].ciede2k)
        elif qty == 'apsnr_y':
            qtys.append(record[key].apsnr_y)
        elif qty == 'apsnr_u':
            qtys.append(record[key].apsnr_u)
        elif qty == 'apsnr_v':
            qtys.append(record[key].apsnr_v)
        elif qty == 'overall_apsnr':
            qtys.append(record[key].overall_apsnr)
        else:
            assert(0)
    return qtys

def CalcBDRate(anchor, test):
    anchor_qty = {}; test_qty = {}
    br_anchor = []; br_test = []
    for key in anchor.keys():
        br_anchor.append(anchor[key].bitrate)
    for key in test.keys():
        br_test.append(test[key].bitrate)

    for qty in qtys:
        anchor_qty[qty]   = GetQty(anchor, qty)
        test_qty[qty] = GetQty(test, qty)

    bdrate = {}; err = 0
    for qty in qtys:
        (err, bdrate[qty]) = BD_RATE(qty, br_anchor, anchor_qty[qty],
                                     br_test, test_qty[qty])
    return bdrate

def CalcFullBDRate(cfg):
    bdrate = {}; seq_time = {}; seq_instr = {}
    for video in records[anchor][cfg].keys():
        bdrate[video] = {}; seq_time[video] = {}
        seq_instr[video] = {}
        for mode in records.keys():
            record = records[mode][cfg][video]
            if mode not in seq_time[video].keys():
                seq_time[video][mode] = 0

            if mode not in seq_instr[video].keys():
                seq_instr[video][mode] = 0

            for key in record.keys():
                seq_time[video][mode] += record[key].enc_time
                seq_instr[video][mode] += record[key].enc_instr
            if mode == anchor:
                continue
            bdrate[video][mode] = CalcBDRate(records[anchor][cfg][video], records[mode][cfg][video])
    return (bdrate, seq_time, seq_instr)


def WriteSummaryXlsFile(bdrate, seq_time, seq_instr, summary):
    csv = open(summary+".csv", "wt")
    csv.write('Video,mode')
    for qty in qtys:
        csv.write(',%s'%qty)
    csv.write(",EncTime(s),EncInstr\n")

    wb = xlsxwriter.Workbook(summary + ".xlsx")
    shts = []

    for mode in csv_files.keys():
        total = {}; avg_bdrate = {}
        sht = wb.add_worksheet(mode)
        shts.append(sht)
        row = 0; col = 0
        sht.write(row, 0, 'Video')
        for qty in qtys:
            col = qtys.index(qty) + 1
            sht.write(row, col, qty)
        sht.write(row, col + 1, "EncTime(s)")
        sht.write(row, col + 2, "EncInstr")
        row = 1
        for video in bdrate.keys():
            csv.write("%s,%s"%(video, mode))
            for qty in qtys:
                if mode == anchor:
                    csv.write(',0.0')
                elif not str(bdrate[video][mode][qty]).startswith('Error'):
                    csv.write(',%f' % bdrate[video][mode][qty])
                else:
                    csv.write(',%s' % bdrate[video][mode][qty])

            csv.write(",%f,%f\n" % (seq_time[video][mode], seq_instr[video][mode]))

            sht.write(row, 0, video)
            for qty in qtys:
                if not qty in total.keys():
                    total[qty] = 0
                    avg_bdrate[qty] = 0

                if mode != anchor and not str(bdrate[video][mode][qty]).startswith('Error'):
                    total[qty] += 1
                    avg_bdrate[qty] += bdrate[video][mode][qty]

                col = qtys.index(qty) + 1
                if mode == anchor:
                    sht.write(row, col, 0.00)
                elif (mode in bdrate[video].keys()):
                    sht.write(row, col, bdrate[video][mode][qty])
            sht.write(row, col + 1, seq_time[video][mode])
            sht.write(row, col + 2, seq_instr[video][mode])
            row += 1
        row += 1
        for qty in qtys:
            if total[qty] != 0:
                avg_bdrate[qty] /= total[qty]
            else:
                avg_bdrate[qty] = 0.0

        sht.write(row, 0, "Average")
        for qty in qtys:
            col = qtys.index(qty) + 1
            sht.write(row, col, avg_bdrate[qty])
        row += 1

    wb.close()
    csv.close()

######################################
# main
######################################
if __name__ == "__main__":
    records = {}
    for tag in csv_files.keys():
        records[tag] = {}
        for test_cfg in csv_files[tag].keys():
            records[tag][test_cfg] = ParseCSVFile(csv_files[tag][test_cfg])

    FillXlsFile()
    DrawRDCurve(records, anchor, rd_curve_pdf)

    #Calculate BDRate and collect total time
    for test_cfg in csv_files[anchor].keys():
        (bdrate, seq_time, seq_instr) = CalcFullBDRate(test_cfg)
        #Write output summary xls file
        filename = "Summary-HEVC-AV1-%s"%test_cfg
        WriteSummaryXlsFile(bdrate, seq_time, seq_instr, filename)
